from google.cloud import api_keys_v2
from google.cloud.api_keys_v2 import Key

from google.cloud import firestore
from config import project_id, api

from google.cloud import firestore


def add_api_key_to_firestore(project_id: str, key: str, user_id: str, key_id: str, name: str) -> None:
    client = firestore.Client(project=project_id, database="aireader")
    doc_ref = client.collection('ApiKeys').document(key)
    doc_ref.set({
        'key': key,
        'user_id': user_id,
        'key_id': key_id,
        'name': name
    })


def create_api_key(project_id: str, id: str, name: str) -> Key:
    """
    Creates and restrict an API key. Add the suffix for uniqueness.

    TODO(Developer):
    1. Before running this sample,
      set up ADC as described in https://cloud.google.com/docs/authentication/external/set-up-adc
    2. Make sure you have the necessary permission to create API keys.

    Args:
        project_id: Google Cloud project id.

    Returns:
        response: Returns the created API Key.
    """
    # Create the API Keys client.
    client = api_keys_v2.ApiKeysClient()

    key = api_keys_v2.Key()
    key.display_name = name

    # Initialize request and set arguments.
    request = api_keys_v2.CreateKeyRequest()
    request.parent = f"projects/{project_id}/locations/global"
    request.key = key
    request.key_id = id

    # Make the request and wait for the operation to complete.
    response = client.create_key(request=request).result()

    print(f"Successfully created an API key: {response.name}")
    # For authenticating with the API key, use the value in "response.key_string".
    # To restrict the usage of this API key, use the value in "response.name".
    return response


def restrict_api_key_api(project_id: str, service: str, key_id: str) -> Key:
    """
    Restricts an API key. Restrictions specify which APIs can be called using the API key.

    TODO(Developer): Replace the variables before running the sample.

    Args:
        project_id: Google Cloud project id.
        key_id: ID of the key to restrict. This ID is auto-created during key creation.
            This is different from the key string. To obtain the key_id,
            you can also use the lookup api: client.lookup_key()

    Returns:
        response: Returns the updated API Key.
    """

    # Create the API Keys client.
    client = api_keys_v2.ApiKeysClient()

    # Restrict the API key usage by specifying the target service and methods.
    # The API key can only be used to authenticate the specified methods in the service.
    api_target = api_keys_v2.ApiTarget()
    api_target.service = service
    api_target.methods = ["*"]

    # Set the API restriction(s).
    # For more information on API key restriction, see:
    # https://cloud.google.com/docs/authentication/api-keys
    restrictions = api_keys_v2.Restrictions()
    restrictions.api_targets = [api_target]

    key = api_keys_v2.Key()
    key.name = f"projects/{project_id}/locations/global/keys/{key_id}"
    key.restrictions = restrictions

    # Initialize request and set arguments.
    request = api_keys_v2.UpdateKeyRequest()
    request.key = key
    request.update_mask = "restrictions"

    # Make the request and wait for the operation to complete.
    response = client.update_key(request=request).result()

    print(f"Successfully updated the API key: {response.name}")
    # Use response.key_string to authenticate.
    return response


def get_students_from_excel() -> list:
    from openpyxl import load_workbook
    wb = load_workbook(filename='Namelist.xlsx')
    sheet = wb.active
    students = []
    for row in sheet.iter_rows(min_row=2, max_col=2):
        id = ""
        name = ""
        for cell in row:
            if cell.column == 1:
                id = cell.value
            if cell.column == 2:
                name = cell.value
        students.append({"id": id, "name": name})
    return students


def get_all_api_keys():
    client = firestore.Client(project=project_id, database="aireader")
    api_keys_ref = client.collection("ApiKeys")
    query = api_keys_ref.order_by("user_id")
    results = [doc.to_dict() for doc in query.stream()]
    print(results)
    return results


if __name__ == "__main__":
    # user_id = "123456789"
    # key = create_api_key(project_id, "studentid-" + user_id ,"cywong@vtc.edu.hk")
    # print(key)
    # response = restrict_api_key_api(project_id, api, key.uid)
    # print(response)
    # add_api_key_to_firestore(project_id, key.key_string, user_id, key.uid)

    existing_keys = get_all_api_keys()
    existing_user_ids = list(map(lambda x: str(x["user_id"]),existing_keys))

    students = get_students_from_excel()
    for student in students:
        # if student["id"] in existing_user_ids skip it
        if student["id"] is None or str(student["id"]) in existing_user_ids:
            continue
        key = create_api_key(project_id, "studentid-" + str(student["id"]) ,student["name"])
        response = restrict_api_key_api(project_id, api, key.uid)
        add_api_key_to_firestore(project_id, key.key_string, student["id"], key.uid, student["name"])
        print(key)
    print("done")
